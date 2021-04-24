#!/usr/bin/env python
from openpyxl import load_workbook
import os
path_in='public_html/prices/filtered'
path_in_arc='public_html/prices/filtered/in_arc'
path_out='public_html/prices/filtered/out'

def convert_file(file):
    current_path=( os.path.dirname (os.path.abspath(__file__)) )
    csv_file =file.replace('xlsx','csv')
    f2=open(  os.path.join(current_path, path_out,csv_file,)  ,'w' )
    wb = load_workbook(filename = os.path.join(current_path,path_in ,file ) )
    sheet_ranges = wb[ wb.sheetnames[0]  ]
    values= sheet_ranges.values
    
    for row in  values:
       csv=[]
       for col in row:
          print str(type(col))
          if str(type(col)) =="<type 'NoneType'>":
              col=''
          csv.append(col.encode('utf8'))
       print  csv[0]
       my_string = ','.join(map(str, csv))
       f2.write( my_string+'\n')
    f2.close()
    wb.close()
    os.rename( os.path.join(current_path,path_in ,file ), os.path.join(current_path,path_in_arc ,file )  )
def main():
    files=[]
    current_path=( os.path.dirname (os.path.abspath(__file__)) )
    if len ( os.listdir(path_in)  )>0:
        for file_name in os.listdir(path_in):
            csv=''
            if file_name.endswith('.xlsx'):
                convert_file(file_name)


if __name__ == "__main__":
    main()
